"""
MedExtractor — HealthWatch Drug Usage Analysis PDF Extractor
============================================================
Extracts all medications from a HealthWatch PDF (image-based/scanned),
then enriches each with Ontario clinical and coverage data via OpenAI.

Output: Colour-coded Excel file with ODB status, brand names, patient programs.

Usage:
    python extract.py fast_movers.pdf
    python extract.py fast_movers.pdf --output results.xlsx
    python extract.py fast_movers.pdf --no-enrich   (extract only, skip AI)
"""

import re, sys, os, json, time, argparse, base64
from pathlib import Path
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()
OPENAI_KEY = os.getenv("OPENAI_API_KEY")

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def ensure_packages():
    pkgs = {"pdfplumber":"pdfplumber","pdf2image":"pdf2image",
            "pytesseract":"pytesseract","openpyxl":"openpyxl","openai":"openai"}
    for import_name, pkg in pkgs.items():
        try:
            __import__(import_name)
        except ImportError:
            print(f"Installing {pkg}...")
            os.system(f"{sys.executable} -m pip install {pkg} --quiet")

ensure_packages()

from pdf2image import convert_from_path
import pytesseract


# ─── STEP 1: OCR EXTRACTION ───────────────────────────────────────────────────

def ocr_pdf(pdf_path: str, dpi: int = 220) -> str:
    """Convert every page to image and OCR it, return full text."""
    print(f"  Converting PDF to images (dpi={dpi})...")
    pages = convert_from_path(pdf_path, dpi=dpi)
    print(f"  OCR-ing {len(pages)} pages...")
    all_text = []
    for i, page in enumerate(pages):
        text = pytesseract.image_to_string(page, config="--psm 6")
        all_text.append(text)
    return "\n\n--- PAGE BREAK ---\n\n".join(all_text)


# ─── STEP 2: PARSE DRUG BLOCKS ────────────────────────────────────────────────

# Matches drug name lines: ALL CAPS words + strength number + unit
DRUG_LINE_RE = re.compile(
    r'^([A-Z][A-Z0-9\-/\.\s]+(?:\d+(?:\.\d+)?(?:\s?(?:MG|MCG|G|ML|%|IU|SENSOR|CRM|HFA|NAS SPR|CTD|SPR|PATCH))[A-Z0-9\s/\.%]*))$'
)
UPC_RE = re.compile(r'^\d{9,14}$')
DIN_RE = re.compile(r'^\d{8}$')
SKIP   = {
    'Drug Usage Analysis','HealthWatch','HealthWaTcH','TONY HUYNH','BROOKLIN',
    'Phone:','Drug Name','UPC','DIN','Margin','Cash','Amount','Page:',
    'Printed:','# Txs','Qty','AAC','Pack','Mfr','BALDWIN','5979',
    'Totals','TP','PAGE BREAK','---','#Txs'
}

def is_skip(line: str) -> bool:
    return any(s in line for s in SKIP)

def parse_strength(name_full: str):
    m = re.search(
        r'(\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)?\s*(?:MG|MCG|G\b|ML\b|%|IU|SENSOR|CRM|HFA|NAS\s*SPR|CTD|SPR)[\w\s/\.%]*)',
        name_full, re.IGNORECASE
    )
    if m:
        strength = m.group(1).strip()
        name = name_full[:m.start()].strip()
        return name, strength
    return name_full.strip(), ''

def parse_drugs(full_text: str) -> list[dict]:
    drugs = []
    current = None
    din_seen = False

    for raw_line in full_text.split('\n'):
        line = raw_line.strip()
        if not line or is_skip(line):
            continue

        parts = line.split()
        first = parts[0] if parts else ''

        # Drug name: ALL CAPS, contains a strength
        is_drug = (
            line == line.upper() and
            re.search(r'\d+(?:\.\d+)?(?:/\d+)?\s*(?:MG|MCG|%|G\b|ML\b|IU|SENSOR)', line, re.I) and
            not UPC_RE.match(first) and not DIN_RE.match(first) and
            len(line) > 5
        )

        if is_drug:
            if current:
                drugs.append(current)
            name, strength = parse_strength(line)
            current = {
                'full_name': line, 'name': name, 'strength': strength,
                'upc': '', 'din': '', 'pack': '', 'manufacturer_code': '',
                'margin_dollar': '', 'margin_pct': ''
            }
            din_seen = False
            continue

        if current is None:
            continue

        # UPC (9–14 digits)
        if UPC_RE.match(first) and not current['upc']:
            current['upc'] = first
            # Pack size is often on same line after UPC
            if len(parts) >= 2 and re.match(r'^\d+\.?\d*$', parts[1]):
                current['pack'] = parts[1]
            continue

        # DIN (8 digits) - comes after UPC
        if DIN_RE.match(first) and not din_seen:
            current['din'] = first
            din_seen = True
            if len(parts) >= 2 and not re.match(r'^\d', parts[1]):
                current['manufacturer_code'] = parts[1]
            elif len(parts) >= 3:
                current['pack'] = parts[1] if re.match(r'^\d', parts[1]) else current['pack']
                current['manufacturer_code'] = parts[2] if len(parts) > 2 else ''
            # Margin dollar (e.g. 1678.05)
            md = re.search(r'\b(\d{1,5}\.\d{2})\b', line)
            if md:
                current['margin_dollar'] = md.group(1)
            # Margin %
            mp = re.search(r'(\d{1,3})\.\d{2}%', line)
            if mp:
                current['margin_pct'] = mp.group(0)
            continue

    if current:
        drugs.append(current)

    # Deduplicate by DIN (prefer entries with DIN)
    seen, unique = set(), []
    for d in drugs:
        key = d['din'] if d['din'] else d['full_name']
        if key not in seen and d['din']:  # only keep if DIN found
            seen.add(key)
            unique.append(d)

    return unique


# ─── STEP 3: OPENAI ENRICHMENT ────────────────────────────────────────────────

ENRICH_PROMPT = """You are an expert Canadian pharmacist. For the medication below, return ONLY valid JSON.

{{
  "disease": "Primary condition(s) treated — be specific (e.g. Type 2 Diabetes, not just Diabetes)",
  "drug_class": "Pharmacological class (e.g. SGLT2 Inhibitor, Statin, Beta-blocker)",
  "brand_names": ["Brand1", "Brand2"],
  "is_generic": true,
  "odb_status": "Covered | Limited Use | Not Covered | Unknown",
  "odb_notes": "LU code if applicable, or 'General benefit' or specific criteria. Be concise.",
  "patient_assistance": [
    {{"program": "Name", "provider": "Company", "benefit": "What it covers/saves", "url": "URL or empty"}}
  ],
  "copay_card": true,
  "ontario_benefits": "Other Ontario programs (OHIP+, Trillium, etc.) or 'None'",
  "notes": "1-2 sentence clinical note for a pharmacist — interactions, monitoring, dispensing tips"
}}

Drug: {name}
Strength: {strength}
DIN: {din}
Manufacturer Code: {mfr}

ODB context for Ontario:
- Most generic chronic disease drugs (statins, ACE inhibitors, beta blockers, metformin) are ODB Covered
- Apixaban (Eliquis) = ODB Limited Use (LU code: AF, DVT, PE, hip/knee replacement)
- Jardiance (empagliflozin) = ODB Covered for T2DM + cardiovascular risk
- FreeStyle Libre = ODB covered for insulin-dependent diabetes
- Dayvigo (lemborexant) = Not covered ODB (newer sleep drug)
- Jublia (efinaconazole) = ODB Limited Use
- Estrogel = ODB Covered
Return only JSON, no explanation."""


def enrich_drug(drug: dict, client) -> dict:
    try:
        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role":"user","content":ENRICH_PROMPT.format(
                name=drug['name'], strength=drug['strength'],
                din=drug['din'], mfr=drug['manufacturer_code']
            )}],
            temperature=0.1, max_tokens=700
        )
        raw = resp.choices[0].message.content.strip()
        m = re.search(r'\{[\s\S]*\}', raw)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"    ⚠ {drug['name']}: {e}")
    return {}


def enrich_all(drugs: list[dict]) -> list[dict]:
    if not OPENAI_KEY:
        print("⚠ No OPENAI_API_KEY in .env — skipping enrichment.")
        return drugs

    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_KEY)
    total = len(drugs)

    for i, drug in enumerate(drugs):
        print(f"  [{i+1}/{total}] {drug['full_name']}")
        info = enrich_drug(drug, client)
        drug.update({
            'disease':           info.get('disease', ''),
            'drug_class':        info.get('drug_class', ''),
            'brand_names':       ', '.join(info.get('brand_names', [])),
            'is_generic':        'Yes' if info.get('is_generic') else 'No',
            'odb_status':        info.get('odb_status', 'Unknown'),
            'odb_notes':         info.get('odb_notes', ''),
            'patient_assistance':'; '.join(
                f"{p.get('program','')} — {p.get('benefit','')}"
                for p in info.get('patient_assistance', [])
            ),
            'copay_card':        'Yes ⭐' if info.get('copay_card') else 'No',
            'ontario_benefits':  info.get('ontario_benefits', ''),
            'ai_notes':          info.get('notes', ''),
        })
        time.sleep(0.6)

    return drugs


# ─── STEP 4: EXCEL OUTPUT ─────────────────────────────────────────────────────

def save_excel(drugs: list[dict], output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Medications"
    ws.sheet_view.showGridLines = False

    FILL_HEADER  = PatternFill("solid", start_color="1F4E79")
    FILL_SUBHEAD = PatternFill("solid", start_color="2E75B6")
    FILL_GREEN   = PatternFill("solid", start_color="E2EFDA")
    FILL_YELLOW  = PatternFill("solid", start_color="FFF2CC")
    FILL_RED     = PatternFill("solid", start_color="FCE4D6")
    FILL_ALT     = PatternFill("solid", start_color="F5F7FB")
    FILL_WHITE   = PatternFill("solid", start_color="FFFFFF")
    THIN = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr(cell, val, fill=FILL_SUBHEAD, size=9, color="FFFFFF", bold=True, wrap=True):
        cell.value = val
        cell.font = Font(name="Arial", bold=bold, size=size, color=color)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border = BORDER

    def cell_set(cell, val, fill, bold=False, color="000000", wrap=True, align="left"):
        cell.value = val
        cell.font = Font(name="Arial", size=9, bold=bold, color=color)
        cell.fill = fill
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        cell.border = BORDER

    # ── Title ──
    ws.merge_cells("A1:R1")
    t = ws["A1"]
    t.value = "HealthWatch Drug Usage Analysis — Ontario Medication Intelligence"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = FILL_HEADER
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Legend ──
    ws.merge_cells("A2:R2")
    lg = ws["A2"]
    lg.value = "  🟢 ODB Covered    🟡 Limited Use (LU Code Required)    🔴 Not Covered / Unknown    ⭐ Copay / Patient Assistance Card Available"
    lg.font = Font(name="Arial", italic=True, size=9, color="444444")
    lg.alignment = Alignment(horizontal="left", vertical="center")
    lg.fill = PatternFill("solid", start_color="EEF2F8")
    ws.row_dimensions[2].height = 20

    # ── Column headers ──
    headers = [
        "Drug Name", "Strength", "DIN", "UPC", "Mfr\nCode",
        "Pack\nSize", "Margin $", "Margin %",
        "Drug Class", "Treats / Disease", "Brand Name(s)",
        "Generic?", "ODB Status", "ODB Notes",
        "Patient Assistance Programs", "Copay\nCard?",
        "Ontario Benefits", "Pharmacist Notes"
    ]
    for col, h in enumerate(headers, 1):
        hdr(ws.cell(row=3, column=col), h)
    ws.row_dimensions[3].height = 36

    # ── Data rows ──
    for i, d in enumerate(drugs):
        row = 4 + i
        ws.row_dimensions[row].height = 60
        odb = d.get('odb_status', '')
        if 'Covered' in odb and 'Not' not in odb and 'Limited' not in odb:
            bg = FILL_GREEN
        elif 'Limited' in odb:
            bg = FILL_YELLOW
        elif 'Not Covered' in odb or 'Unknown' in odb:
            bg = FILL_RED
        else:
            bg = FILL_ALT if i % 2 == 0 else FILL_WHITE

        vals = [
            d.get('name','') or d.get('full_name',''),
            d.get('strength',''),
            d.get('din',''),
            d.get('upc',''),
            d.get('manufacturer_code',''),
            d.get('pack',''),
            d.get('margin_dollar',''),
            d.get('margin_pct',''),
            d.get('drug_class',''),
            d.get('disease',''),
            d.get('brand_names',''),
            d.get('is_generic',''),
            d.get('odb_status',''),
            d.get('odb_notes',''),
            d.get('patient_assistance',''),
            d.get('copay_card',''),
            d.get('ontario_benefits',''),
            d.get('ai_notes',''),
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col)
            is_bold = col == 1
            txt_color = "000000"
            # Colour ODB cell
            if col == 13:
                if 'Covered' in str(val) and 'Not' not in str(val) and 'Limited' not in str(val):
                    txt_color = "1A7A1A"
                elif 'Limited' in str(val):
                    txt_color = "7A5A00"
                elif 'Not' in str(val) or 'Unknown' in str(val):
                    txt_color = "C00000"
                is_bold = True
            cell_set(c, val, bg, bold=is_bold, color=txt_color)

    # ── Column widths ──
    widths = [22, 14, 10, 14, 7, 7, 10, 9, 22, 30, 22, 8, 14, 30, 38, 8, 22, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:R{3+len(drugs)}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:C1")
    sh = ws2["A1"]
    sh.value = "ODB Coverage Summary"
    sh.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    sh.fill = FILL_HEADER
    sh.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col, label in enumerate(["ODB Status","Count","% of Total"],1):
        hdr(ws2.cell(row=2,column=col), label)

    odb_counts = {}
    copay_total = 0
    for d in drugs:
        s = d.get('odb_status','Unknown')
        odb_counts[s] = odb_counts.get(s,0) + 1
        if 'Yes' in d.get('copay_card',''):
            copay_total += 1

    r = 3
    for status, cnt in sorted(odb_counts.items()):
        ws2.cell(row=r,column=1).value = status
        ws2.cell(row=r,column=2).value = cnt
        pc = ws2.cell(row=r,column=3)
        pc.value = cnt/len(drugs) if drugs else 0
        pc.number_format = "0.0%"
        for col in range(1,4):
            ws2.cell(row=r,column=col).font = Font(name="Arial",size=9)
            ws2.cell(row=r,column=col).border = BORDER
        r+=1

    ws2.cell(row=r,column=1).value = "Total Drugs"
    ws2.cell(row=r,column=2).value = len(drugs)
    ws2.cell(row=r+1,column=1).value = "With Copay Card ⭐"
    ws2.cell(row=r+1,column=2).value = copay_total
    for rr in [r,r+1]:
        for col in [1,2]:
            ws2.cell(row=rr,column=col).font = Font(name="Arial",bold=True,size=9)
            ws2.cell(row=rr,column=col).border = BORDER

    for col,w in [(1,22),(2,10),(3,12)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"\n✅ Saved → {output_path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Extract medications from HealthWatch PDF")
    ap.add_argument('pdf', help='Path to the HealthWatch PDF')
    ap.add_argument('--output','-o', default='medication_analysis.xlsx')
    ap.add_argument('--no-enrich', action='store_true', help='Skip OpenAI enrichment')
    ap.add_argument('--dpi', type=int, default=220, help='OCR resolution (default 220)')
    args = ap.parse_args()

    if not Path(args.pdf).exists():
        print(f"❌ File not found: {args.pdf}"); sys.exit(1)

    print(f"\n⚕  MedExtractor")
    print(f"   PDF    : {args.pdf}")
    print(f"   Output : {args.output}")
    print("─"*50)

    print("\n📄 Step 1: OCR — reading PDF pages...")
    full_text = ocr_pdf(args.pdf, dpi=args.dpi)

    print("\n🔍 Step 2: Parsing drug records...")
    drugs = parse_drugs(full_text)
    print(f"   Found {len(drugs)} medications")

    if not drugs:
        print("⚠ No drugs parsed. Check PDF format.")
        sys.exit(1)

    if not args.no_enrich:
        print(f"\n🧠 Step 3: AI enrichment via OpenAI ({len(drugs)} drugs)...")
        drugs = enrich_all(drugs)
    else:
        print("\n⏩ Skipping AI enrichment.")

    print("\n📊 Step 4: Building Excel report...")
    save_excel(drugs, args.output)

    covered = sum(1 for d in drugs if 'Covered' in d.get('odb_status','') and 'Not' not in d.get('odb_status','') and 'Limited' not in d.get('odb_status',''))
    limited = sum(1 for d in drugs if 'Limited' in d.get('odb_status',''))
    cards   = sum(1 for d in drugs if 'Yes' in d.get('copay_card',''))

    print(f"\n📋 Results:")
    print(f"   Drugs extracted  : {len(drugs)}")
    print(f"   ODB Covered      : {covered}")
    print(f"   ODB Limited Use  : {limited}")
    print(f"   Copay Cards      : {cards}")
    print(f"\n   → {args.output}\n")


if __name__ == "__main__":
    main()
